# Васю назначили завхозом в туристической группе и он подошёл к подготовке
# ответственно, составив справочник продуктов с указанием калорийности на 100
# грамм, а также содержание белков, жиров и углеводов на 100 грамм продукта.
# Ему не удалось найти всю информацию, поэтому некоторые ячейки остались
# незаполненными (можно считать их значение равным нулю). Также он использовал
# какой-то странный офисный пакет и разделял целую и дробную часть чисел
# запятой. Таблица доступа по ссылке
# https://stepik.org/media/attachments/lesson/245290/trekking1.xlsx
#
# Вася хочет минимизировать вес продуктов и для этого брать самые калорийные
# продукты. Помогите ему и упорядочите продукты по убыванию калорийности. В
# случае, если продукты имеют одинаковую калорийность - упорядочите их по
# названию. В качестве ответа необходимо сдать названия продуктов, по одному в
# строке.
from openpyxl import load_workbook


def fix(s):
    if s is None:
        return 0.

    s = str(s)
    if ',' in s:
        s.replace(',', '.')

    return float(s)

wb = load_workbook('input_w06e15.xlsx')
sh = wb['Справочник']

m_row = sh.max_row

calories = {}

for i in range(2, m_row+1):

    prod = sh.cell(row=i, column=1).value
    val = fix(sh.cell(row=i, column=2).value)

    if val in calories:
        calories[val].append(prod)
    else:
        calories[val] = [prod]

sorted_calories = sorted(calories, reverse=True)

for c in sorted_calories:
    for p in sorted(calories[c]):
        print(p)
