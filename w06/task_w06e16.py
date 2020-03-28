# Васю назначили завхозом в туристической группе и он подошёл к подготовке
# ответственно, составив справочник продуктов с указанием калорийности на 100
# грамм, а также содержание белков, жиров и углеводов на 100 грамм продукта.
# Ему не удалось найти всю информацию, поэтому некоторые ячейки остались
# незаполненными (можно считать их значение равным нулю). Также он использовал
# какой-то странный офисный пакет и разделял целую и дробную часть чисел
# запятой. Таблица доступна по ссылке
# https://stepik.org/media/attachments/lesson/245290/trekking2.xlsx
#
# Вася составил раскладку по продуктам на один день (она на листе "Раскладка")
# с указанием названия продукта и его количества в граммах. Посчитайте 4 числа:
# суммарную калорийность и граммы белков, жиров и углеводов. Числа округлите до
# целых вниз и введите через пробел.

from openpyxl import load_workbook


def fix(s):
    if s is None:
        return 0.

    s = str(s)
    if ',' in s:
        s.replace(',', '.')

    return float(s)


def raskladka(sheet):

    res = []

    m_row = sheet.max_row

    for i in range(2, m_row+1):
        key = sheet.cell(row=i, column=1).value
        value = fix(sheet.cell(row=i, column=2).value)
        res.append([])
        res[i-2].append(key)
        res[i-2].append(value)

    return res


def spravochnik(sheet):

    res = {}

    m_row = sheet.max_row

    for i in range(2, m_row + 1):
        key = sheet.cell(row=i, column=1).value
        res[key] = []
        for j in range(2, 6):
            res[key].append(fix(sheet.cell(row=i, column=j).value))

    return res


wb = load_workbook('input_w06e16.xlsx')
sh1 = wb['Справочник']
sh2 = wb['Раскладка']

raskl = raskladka(sh2)
sprav = spravochnik(sh1)

calories = 0
proteins = 0
fats = 0
carbohydrates = 0

for i in range(len(raskl)):
    prod = raskl[i][0]
    mass = raskl[i][1]
    calories += sprav[prod][0] * mass /100
    proteins += sprav[prod][1] * mass / 100
    fats += sprav[prod][2] * mass / 100
    carbohydrates += sprav[prod][3] * mass / 100

print(int(calories), int(proteins), int(fats), int(carbohydrates))