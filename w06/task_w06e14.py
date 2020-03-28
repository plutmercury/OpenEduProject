# Вася планирует карьеру и переезд. Для это составил таблицу, в которой для
# каждого региона записал зарплаты для разных интересные ему профессий.
# Таблица доступна по ссылке
# https://stepik.org/media/attachments/lesson/245267/salaries.xlsx.
# Выведите название региона с самой высокой медианной зарплатой (медианой
# называется элемент, стоящий в середине массива после его упорядочивания)
# и, через пробел, название профессии с самой высокой средней зарплатой по всем
# регионам.

def average(lst):
    if len(lst) > 0:
        return sum(lst)/len(lst)
    return 0


from openpyxl import load_workbook

wb = load_workbook('input_w06e14.xlsx')

sheet = wb['Лист1']
sheet_rows = sheet.max_row
sheet_columns = sheet.max_column

max_median = 0
max_median_region = ''
for i in range(2, sheet_rows+1):
    salary_in_region = []
    for j in range(2, sheet_columns+1):
        salary_in_region.append(sheet.cell(row=i, column=j).value)
    if max_median < sorted(salary_in_region)[(sheet_columns-1) // 2]:
        max_median = sorted(salary_in_region)[(sheet_columns-1) // 2]
        max_median_region = sheet.cell(row=i, column=1).value

max_average_sallary = 0
max_average_sallary_position =''
for j in range(2, sheet_columns+1):
    salary_for_position = []
    for i in range(2, sheet_rows+1):
        salary_for_position.append(sheet.cell(row=i, column=j).value)
    if max_average_sallary < average(salary_for_position):
        max_average_sallary = average(salary_for_position)
        max_average_sallary_position = sheet.cell(row=1, column=j).value

print(max_median_region, max_average_sallary_position)

