# Главный бухгалтер компании "Рога и копыта" случайно удалил ведомость с
# начисленной зарплатой. К счастью, у него сохранились расчётные листки всех
# сотрудников. Помогите по этим расчётным листкам восстановить зарплатную
# ведомость. Архив с расчётными листками доступен по ссылке
# https://stepik.org/media/attachments/lesson/245299/rogaikopyta.zip (вы
# можете скачать и распаковать его вручную или самостоятельно научиться делать
# это с помощью скрипта на Питоне).
#
# Ведомость должна содержать 1000 строк, в каждой строке должно быть указано
# ФИО сотрудника и, через пробел, его зарплата. Сотрудники должны быть
# упорядочены по алфавиту.

from openpyxl import load_workbook



vedomost = {}
for i in range(1, 1001):
    rl = load_workbook('input_w06e13/' + str(i) + '.xlsx')
    vedomost[rl['Sheet'].cell(row=2, column=2).value] = rl['Sheet'].cell(row=2, column=4).value

arr_vedomost = sorted(vedomost)

fout = open('output_w06e13.txt', 'w', encoding='utf8')

for key in arr_vedomost:
    print(key, vedomost[key], file=fout)

fout.close()