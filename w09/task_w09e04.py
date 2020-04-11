# Вам дан xlsx-файл с информацией о спортивных площадках в Москве
# https://stepik.org/media/attachments/lesson/268674/data-25290-2019-09-30.xlsx.
# Вам необходимо сдать на проверку json-файл, в котором будет храниться один
# словарь, ключами которого будут административные округа (AdmArea), а
# значениями словари, в которых, в свою очередь, ключами будут названия районов
# (District), относящихся к этому административному округу, а значениями -
# списки адресов площадок (Address) в том порядке, в котором они встречались в
# исходном файле.
#
# Ваш файл должен выглядеть примерно так:
#
# { "Северо-Западный административный округ": {
#   "район Строгино": [
#       "улица Исаковского, дом 24, корпус 1",
#       "Неманский проезд, дом 9"
#    ],
#   "район Северное Тушино": [
#       "улица Свободы, дом 56",
#       "улица Свободы, дом 56",
#       "улица Свободы, дом 56",
#       "улица Свободы, дом 56"
#   ],
#   "район Покровское-Стрешнево": [
#       "Иваньковское шоссе, дом 6"
#   ],
#   "район Щукино": [
#   "Сосновая улица, дом 3, строение 2"
#   ]
# }}

from openpyxl import load_workbook
import json

wb = load_workbook('input_w09e04.xlsx')
sheet = wb['Sheet0']

result = {}

max_row = sheet.max_row + 1
for i in range(2, max_row):
    admin_area = sheet.cell(row=i, column=5).value
    district = sheet.cell(row=i, column=6).value
    address = sheet.cell(row=i, column=7).value

    if not admin_area in result:
        result[admin_area] = {}
    if not district in result[admin_area]:
        result[admin_area][district] = []
    result[admin_area][district].append(address)

fout = open('output_w09e04.json', 'w', encoding='utf8')
json.dump(result, fout, ensure_ascii=False)
fout.close()